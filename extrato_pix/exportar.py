# -*- coding: utf-8 -*-
"""
Exportação do resultado para: .xlsx (Excel), .pdf, .csv ou .txt.

O formato é escolhido automaticamente pela extensão do arquivo que o usuário
digitar ao salvar (veja a função 'exportar').
"""

import csv
import os

from .valores import formatar_brl, formatar_numero_br


# ===========================================================================
# CSV (separador ';', compatível com Excel em português)
# ===========================================================================
def exportar_csv(resultado, caminho):
    """Salva o resultado em CSV (separador ';')."""
    # utf-8-sig faz o Excel reconhecer os acentos corretamente.
    with open(caminho, "w", newline="", encoding="utf-8-sig") as arquivo:
        escritor = csv.writer(arquivo, delimiter=";")
        escritor.writerow(["#", "Descrição", "Valor (R$)"])
        for indice, transacao in enumerate(resultado.transacoes, start=1):
            escritor.writerow(
                [indice, transacao.descricao, formatar_numero_br(transacao.valor)]
            )
        escritor.writerow([])
        escritor.writerow(["", "TOTAL", formatar_numero_br(resultado.total)])
        escritor.writerow(["", "Quantidade de transações", resultado.quantidade])


# ===========================================================================
# Excel .xlsx (bem formatado, com colunas separadas e valores como números)
# ===========================================================================
def exportar_xlsx(resultado, caminho):
    """Salva o resultado em uma planilha Excel (.xlsx) bonita e organizada."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    # Formato de moeda do Excel. Os separadores (milhar/decimal) seguem o
    # idioma do Windows — em português aparece como  R$ 1.234,56.
    FORMATO_MOEDA = '"R$" #,##0.00'
    VERDE = "0F766E"
    VERDE_CLARO = "E6F4F1"

    wb = Workbook()
    ws = wb.active
    ws.title = "PIX recebidos"

    # ---- Cabeçalho do relatório ----
    ws["A1"] = "Doces da Praia — PIX recebidos"
    ws["A1"].font = Font(size=15, bold=True, color=VERDE)
    ws["A2"] = f"Arquivo: {os.path.basename(resultado.arquivo)}"
    ws["A3"] = f"Transações encontradas: {resultado.quantidade}"
    ws["A4"] = "Total:"
    ws["B4"] = round(resultado.total, 2)
    ws["B4"].number_format = FORMATO_MOEDA
    ws["A4"].font = Font(bold=True, color=VERDE)
    ws["B4"].font = Font(bold=True, color=VERDE)

    # ---- Cabeçalho da tabela ----
    linha_cab = 6
    titulos = ["#", "Descrição", "Valor (R$)"]
    borda_fina = Side(style="thin", color="D5DBDB")
    for coluna, titulo in enumerate(titulos, start=1):
        celula = ws.cell(row=linha_cab, column=coluna, value=titulo)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor=VERDE)
        celula.alignment = Alignment(horizontal="center", vertical="center")
        celula.border = Border(bottom=borda_fina)

    # ---- Linhas de dados ----
    linha = linha_cab + 1
    for indice, transacao in enumerate(resultado.transacoes, start=1):
        ws.cell(row=linha, column=1, value=indice).alignment = Alignment(horizontal="center")
        ws.cell(row=linha, column=2, value=transacao.descricao)
        celula_valor = ws.cell(row=linha, column=3, value=round(transacao.valor, 2))
        celula_valor.number_format = FORMATO_MOEDA

        # Destaca em vermelho as linhas cujo valor não foi reconhecido.
        if not transacao.valor_encontrado:
            celula_valor.font = Font(color="C0392B")

        # Faixa zebrada (linhas alternadas) para facilitar a leitura.
        if indice % 2 == 0:
            for coluna in range(1, 4):
                ws.cell(row=linha, column=coluna).fill = PatternFill("solid", fgColor=VERDE_CLARO)
        linha += 1

    # ---- Linha de TOTAL ----
    ws.cell(row=linha, column=2, value="TOTAL").font = Font(bold=True)
    celula_total = ws.cell(row=linha, column=3, value=round(resultado.total, 2))
    celula_total.font = Font(bold=True)
    celula_total.number_format = FORMATO_MOEDA
    celula_total.fill = PatternFill("solid", fgColor=VERDE_CLARO)
    ws.cell(row=linha, column=2).fill = PatternFill("solid", fgColor=VERDE_CLARO)

    # ---- Larguras das colunas + travar o cabeçalho ----
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 72
    ws.column_dimensions["C"].width = 16
    ws.freeze_panes = ws.cell(row=linha_cab + 1, column=1)

    wb.save(caminho)


# ===========================================================================
# PDF (relatório pronto para imprimir)
# ===========================================================================
def _ajustar_texto(pdf, texto, largura_mm):
    """Corta o texto com '...' caso ele não caiba na largura informada."""
    if pdf.get_string_width(texto) <= largura_mm:
        return texto
    while texto and pdf.get_string_width(texto + "...") > largura_mm:
        texto = texto[:-1]
    return texto + "..."


def exportar_pdf(resultado, caminho):
    """Salva o resultado em um arquivo PDF formatado."""
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos

    def latin(texto):
        # As fontes padrão do PDF usam latin-1; troca o que não for suportado.
        return str(texto).encode("latin-1", "replace").decode("latin-1")

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # ---- Título e resumo ----
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(15, 118, 110)
    pdf.cell(0, 10, latin("Doces da Praia - PIX recebidos"),
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(40, 40, 40)
    pdf.cell(0, 6, latin(f"Arquivo: {os.path.basename(resultado.arquivo)}"),
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 6, latin(f"Transacoes encontradas: {resultado.quantidade}"),
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.set_font("Helvetica", "B", 13)
    pdf.set_text_color(15, 118, 110)
    pdf.cell(0, 9, latin(f"TOTAL: {formatar_brl(resultado.total)}"),
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(3)

    # ---- Cabeçalho da tabela ----
    largura_util = 190.0          # 210mm (A4) - 2 x 10mm de margem
    largura_valor = 40.0
    largura_desc = largura_util - largura_valor

    pdf.set_fill_color(15, 118, 110)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(largura_desc, 8, latin("  Descricao"), fill=True,
             new_x=XPos.RIGHT, new_y=YPos.TOP)
    pdf.cell(largura_valor, 8, latin("Valor (R$)  "), fill=True, align="R",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    # ---- Linhas ----
    pdf.set_font("Helvetica", "", 10)
    zebra = False
    for transacao in resultado.transacoes:
        if zebra:
            pdf.set_fill_color(244, 246, 248)
        else:
            pdf.set_fill_color(255, 255, 255)
        zebra = not zebra

        descricao = "  " + _ajustar_texto(pdf, latin(transacao.descricao), largura_desc - 4)
        pdf.set_text_color(20, 20, 20)
        pdf.cell(largura_desc, 7, descricao, fill=True,
                 new_x=XPos.RIGHT, new_y=YPos.TOP)

        # Vermelho quando o valor não foi reconhecido.
        if transacao.valor_encontrado:
            pdf.set_text_color(20, 20, 20)
        else:
            pdf.set_text_color(192, 57, 43)
        pdf.cell(largura_valor, 7, latin(formatar_brl(transacao.valor)) + "  ",
                 fill=True, align="R", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.output(caminho)


# ===========================================================================
# TXT (relatório de texto simples)
# ===========================================================================
def exportar_txt(resultado, caminho):
    """Salva o resultado em um relatório de texto legível (.txt)."""
    largura = 64
    linhas = []
    linhas.append("Doces da Praia - Calcular Extrato")
    linhas.append("Relatório de PIX recebidos")
    linhas.append("=" * largura)
    linhas.append(f"Arquivo analisado : {os.path.basename(resultado.arquivo)}")
    linhas.append(f"Transações        : {resultado.quantidade}")
    linhas.append(f"TOTAL             : {formatar_brl(resultado.total)}")
    linhas.append("=" * largura)
    linhas.append("")
    linhas.append(f"{'VALOR':>15}   DESCRIÇÃO")
    linhas.append("-" * largura)
    for transacao in resultado.transacoes:
        marcador = "" if transacao.valor_encontrado else "  (valor não reconhecido)"
        linhas.append(
            f"{formatar_brl(transacao.valor):>15}   {transacao.descricao}{marcador}"
        )
    with open(caminho, "w", encoding="utf-8") as arquivo:
        arquivo.write("\n".join(linhas))


# ===========================================================================
# Despacho: escolhe o formato pela extensão do arquivo
# ===========================================================================
def exportar(resultado, caminho):
    """Exporta para o formato indicado pela extensão do 'caminho'."""
    extensao = os.path.splitext(caminho)[1].lower()
    if extensao == ".txt":
        exportar_txt(resultado, caminho)
    elif extensao == ".pdf":
        exportar_pdf(resultado, caminho)
    elif extensao in (".xlsx", ".xls"):
        exportar_xlsx(resultado, caminho)
    else:
        exportar_csv(resultado, caminho)
